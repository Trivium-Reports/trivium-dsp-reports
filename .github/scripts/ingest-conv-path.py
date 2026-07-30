#!/usr/bin/env python3
"""
Trivium conversion-path ingest — Gmail API direct, additive to the DSP pipeline.

WHY THIS EXISTS
---------------
The weekly DSP pipeline (refresh-data.py) ingests reports via the auth site's
/gmail-ingest function, which downloads the *pre-signed S3 link* that Amazon DSP
report emails contain. Sponsored Ads "Conversion path" report emails do NOT
contain an S3 link — they only link back into advertising.amazon.com, which
returns Amazon's sign-in page to any unauthenticated fetch. That is why
public/data/<slug>/conv-path.csv has never existed for any brand.

This script closes that gap without touching the working DSP path. It talks to
Gmail directly with the ppc@triviumco.com OAuth credentials and accepts the
report CSV from any of three delivery shapes, in priority order:

  1. ATTACHMENT  — a .csv or .xlsx attached to the email. This is the reliable
     route: download the report once from the Amazon console and forward it to
     ppc@triviumco.com with the brand in the subject. Fully automatic from there.
  2. S3 LINK     — a pre-signed d16g-analytics-reports-*.s3.amazonaws.com URL,
     if Amazon ever ships conv-path reports that way (DSP-console reports do).
  3. CONSOLE LINK — detected and reported as UNFETCHABLE, never written. We do
     not scrape a logged-in session.

Emails are matched on subject: the brand name must appear in it, alongside
"conversion path". Matching is against BRANDS below, longest-name-first so
"Primal Queen Collagen" can't be shadowed by a shorter name.

Writes public/data/<slug>/conv-path.csv. The workflow commits any changes.
Exits 0 even when nothing is found — a missing report must not fail the run and
block the DSP refresh commit.

Env vars (GitHub repo secrets):
  GMAIL_CLIENT_ID, GMAIL_CLIENT_SECRET, GMAIL_REFRESH_TOKEN
"""

import base64
import csv
import io
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

# ── Configuration ────────────────────────────────────────────

# Brands subscribed to a conversion path report. Add a line to onboard a brand;
# nothing else needs changing. "name" must appear in the email subject.
BRANDS = [
    {"slug": "primal-queen", "name": "Primal Queen"},
    {"slug": "mirai-clinical", "name": "Mirai"},
]

LOOKBACK_DAYS = int(os.environ.get("CONV_PATH_LOOKBACK_DAYS", "14"))
DRY_RUN = os.environ.get("CONV_PATH_DRY_RUN", "").lower() in ("1", "true", "yes")

GMAIL_API = "https://gmail.googleapis.com/gmail/v1/users/me"
OUT_ROOT = Path("public/data")

S3_HOST_HINT = "analytics-reports"
CONSOLE_HOST_HINT = "advertising.amazon.com"


# ── HTTP helpers ─────────────────────────────────────────────

def http(url, headers=None, data=None, method=None, timeout=120):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.status, dict(resp.headers), resp.read()
    except urllib.error.HTTPError as e:
        return e.code, dict(e.headers or {}), e.read() if hasattr(e, "read") else b""


def access_token():
    cid = os.environ.get("GMAIL_CLIENT_ID")
    secret = os.environ.get("GMAIL_CLIENT_SECRET")
    refresh = os.environ.get("GMAIL_REFRESH_TOKEN")
    missing = [n for n, v in [("GMAIL_CLIENT_ID", cid),
                              ("GMAIL_CLIENT_SECRET", secret),
                              ("GMAIL_REFRESH_TOKEN", refresh)] if not v]
    if missing:
        sys.exit(f"ERR: missing repo secret(s): {', '.join(missing)}")
    body = urllib.parse.urlencode({
        "client_id": cid,
        "client_secret": secret,
        "refresh_token": refresh,
        "grant_type": "refresh_token",
    }).encode()
    status, _, raw = http("https://oauth2.googleapis.com/token", data=body,
                          headers={"Content-Type": "application/x-www-form-urlencoded"})
    if status != 200:
        sys.exit(f"ERR: token refresh failed ({status}): {raw[:300]!r}")
    return json.loads(raw)["access_token"]


def gmail(path, token, params=None):
    url = f"{GMAIL_API}{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    status, _, raw = http(url, headers={"Authorization": f"Bearer {token}"})
    if status != 200:
        sys.exit(f"ERR: Gmail {path} failed ({status}): {raw[:300]!r}")
    return json.loads(raw)


# ── Message walking ──────────────────────────────────────────

def walk_parts(part):
    yield part
    for child in part.get("parts") or []:
        yield from walk_parts(child)


def header(msg, name):
    for h in msg.get("payload", {}).get("headers", []):
        if h["name"].lower() == name.lower():
            return h["value"]
    return ""


def body_text(msg):
    chunks = []
    for part in walk_parts(msg.get("payload", {})):
        data = part.get("body", {}).get("data")
        if data:
            chunks.append(base64.urlsafe_b64decode(data).decode("utf-8", "replace"))
    return "".join(chunks)


def find_attachment(msg, token):
    """Return (filename, bytes) for the first .csv/.xlsx attachment, else None."""
    for part in walk_parts(msg.get("payload", {})):
        filename = (part.get("filename") or "").strip()
        if not filename.lower().endswith((".csv", ".xlsx")):
            continue
        att_id = part.get("body", {}).get("attachmentId")
        if not att_id:
            continue
        meta = gmail(f"/messages/{msg['id']}/attachments/{att_id}", token)
        return filename, base64.urlsafe_b64decode(meta["data"])
    return None


def find_links(html):
    """Amazon wraps links through na.r.ads.amazon.com with the target URL-encoded."""
    urls = re.findall(r'https?://[^\s"\'<>]+', html)
    decoded = []
    for u in urls:
        decoded.append(u)
        if "r.ads.amazon.com" in u:
            decoded.append(urllib.parse.unquote(u))
    s3 = [u for u in urls if S3_HOST_HINT in u]
    console = [u for u in decoded if "download-report" in u or CONSOLE_HOST_HINT in u]
    return s3, console


# ── Payload → CSV ────────────────────────────────────────────

def looks_like_html(blob):
    head = blob[:2000].lstrip().lower()
    return head.startswith(b"<") or b"<!doctype html" in head or b"<html" in head


def xlsx_to_csv(blob):
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERR: .xlsx attachment needs openpyxl — add it to the workflow's pip install.")
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = io.StringIO()
    writer = csv.writer(out, quoting=csv.QUOTE_ALL)
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        writer.writerow(["" if c is None else c for c in row])
    return out.getvalue()


def to_csv_text(filename, blob):
    """Normalise an attachment or download into CSV text. None if unusable."""
    if looks_like_html(blob):
        return None, "html_not_csv"
    if filename.lower().endswith(".xlsx") or blob[:2] == b"PK":
        text = xlsx_to_csv(blob)
    else:
        text = blob.decode("utf-8-sig", "replace")
    rows = [r for r in text.splitlines() if r.strip()]
    if len(rows) < 2:
        return None, f"header_only_or_empty ({len(rows)} row(s))"
    return text, None


# ── Matching ─────────────────────────────────────────────────

def match_brand(subject):
    subj = subject.lower()
    if "conversion path" not in subj:
        return None
    for brand in sorted(BRANDS, key=lambda b: -len(b["name"])):
        if brand["name"].lower() in subj:
            return brand
    return None


# ── Main ─────────────────────────────────────────────────────

def fetch_report(msg, token):
    """Return (csv_text, source, note). csv_text None means nothing usable."""
    attachment = find_attachment(msg, token)
    if attachment:
        filename, blob = attachment
        text, err = to_csv_text(filename, blob)
        if text:
            return text, f"attachment:{filename}", None
        return None, f"attachment:{filename}", err

    s3_links, console_links = find_links(body_text(msg))
    for link in s3_links:
        status, _, blob = http(link)
        if status != 200:
            continue
        text, err = to_csv_text(link, blob)
        if text:
            return text, "s3_link", None
        return None, "s3_link", err

    if console_links:
        return None, "console_link", (
            "email links into advertising.amazon.com, which needs a signed-in "
            "session — download it once and forward the file to ppc@triviumco.com"
        )
    return None, "none", "no attachment and no download link found"


def main():
    token = access_token()
    # Broad query on purpose: Amazon's subject wording varies ("Conversion path
    # report" vs "Conversion Paths Report"), and Gmail tokenises so a quoted
    # "conversion path" misses the plural. Filter precisely in match_brand().
    query = f'subject:conversion newer_than:{LOOKBACK_DAYS}d'
    listing = gmail("/messages", token, {"q": query, "maxResults": 100})
    messages = listing.get("messages", [])
    print(f"→ Gmail query: {query}")
    print(f"→ {len(messages)} candidate message(s)")

    # Newest first, and only the newest usable report per brand.
    best = {}
    for stub in messages:
        msg = gmail(f"/messages/{stub['id']}", token, {"format": "full"})
        subject = header(msg, "Subject")
        brand = match_brand(subject)
        if not brand:
            print(f"  · skip (no brand match): {subject[:80]}")
            continue
        if brand["slug"] in best:
            continue
        text, source, note = fetch_report(msg, token)
        label = f"{brand['name']} [{source}]"
        if not text:
            print(f"  ✗ {label}: {note}")
            continue
        print(f"  ✓ {label}: {len(text.splitlines())} rows")
        best[brand["slug"]] = text

    if not best:
        print("→ nothing ingestible this run — leaving existing CSVs untouched")
        return 0

    for slug, text in best.items():
        dest = OUT_ROOT / slug / "conv-path.csv"
        if DRY_RUN:
            print(f"→ DRY RUN, would write {dest} ({len(text)} bytes)")
            continue
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(text, encoding="utf-8")
        print(f"→ wrote {dest} ({len(text)} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
